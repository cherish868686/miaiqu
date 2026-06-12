"""
Excel文档解析器 (.xlsx, .xls)
提取每个工作表的表格内容
"""
from openpyxl import load_workbook


def parse_xlsx(filepath):
    """解析Excel文档，返回所有工作表的表格内容"""
    wb = load_workbook(filepath, data_only=True)
    texts = []
    tables = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        table_data = {
            'index': len(tables),
            'sheet_name': sheet_name,
            'headers': [],
            'rows': []
        }

        for r_idx, row in enumerate(ws.iter_rows(values_only=True)):
            row_data = [str(cell) if cell is not None else '' for cell in row]
            if not any(row_data):
                continue
            if r_idx == 0:
                table_data['headers'] = row_data
            else:
                table_data['rows'].append({
                    'row_index': r_idx,
                    'cells': row_data
                })

        tables.append(table_data)

    wb.close()
    return {'texts': texts, 'tables': tables}
